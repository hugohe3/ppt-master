#!/usr/bin/env python3
"""Compile Excel analysis prompt packs and build image-generation manifests."""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from console_encoding import configure_utf8_stdio


configure_utf8_stdio()

REQUIRED_COLUMNS = (
    "item_id",
    "category",
    "name",
    "prompt_template",
    "output_filename",
    "aspect_ratio",
    "image_size",
    "enabled",
    "reference_role",
    "notes",
)
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp"}
DEFAULT_PROFILE = "gptimage2.0-1K-low"


def _require_openpyxl():
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for analysis prompt packs. "
            "Run: pip install openpyxl"
        ) from exc
    return openpyxl


def _slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9-]+", "-", value.strip().lower()).strip("-")
    if not slug:
        raise ValueError("pack_id must contain lowercase letters, digits, or hyphens")
    return slug


def _enabled(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on", "启用", "是"}


def _text(value: Any) -> str:
    return str(value or "").strip()


def compile_workbook(
    workbook_path: str | Path,
    output_path: str | Path | None = None,
    *,
    pack_id: str | None = None,
) -> Path:
    """Validate an Excel prompt pack and compile it to stable JSON."""
    openpyxl = _require_openpyxl()
    source = Path(workbook_path).resolve()
    if not source.is_file():
        raise ValueError(f"Workbook not found: {source}")

    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["prompts"] if "prompts" in workbook.sheetnames else workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [_text(value) for value in next(rows)]
    except StopIteration as exc:
        raise ValueError(f"Workbook is empty: {source}") from exc

    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")
    column_index = {name: headers.index(name) for name in REQUIRED_COLUMNS}

    items: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_outputs: set[str] = set()
    for row_number, values in enumerate(rows, start=2):
        row = {
            name: values[index] if index < len(values) else None
            for name, index in column_index.items()
        }
        if not any(_text(value) for value in row.values()):
            continue
        item_id = _text(row["item_id"])
        if not item_id:
            raise ValueError(f"Row {row_number}: item_id is required")
        if item_id in seen_ids:
            raise ValueError(f"Row {row_number}: duplicate item_id '{item_id}'")
        seen_ids.add(item_id)

        output_filename = _text(row["output_filename"])
        if not output_filename or Path(output_filename).name != output_filename:
            raise ValueError(f"Row {row_number}: output_filename must be a filename")
        if Path(output_filename).suffix.lower() not in IMAGE_SUFFIXES:
            raise ValueError(
                f"Row {row_number}: output_filename must end with "
                f"{', '.join(sorted(IMAGE_SUFFIXES))}"
            )
        if output_filename in seen_outputs:
            raise ValueError(f"Row {row_number}: duplicate output_filename '{output_filename}'")
        seen_outputs.add(output_filename)

        prompt = _text(row["prompt_template"])
        if not prompt:
            raise ValueError(f"Row {row_number}: prompt_template is required")
        aspect_ratio = _text(row["aspect_ratio"])
        image_size = _text(row["image_size"])
        if not aspect_ratio or not image_size:
            raise ValueError(f"Row {row_number}: aspect_ratio and image_size are required")

        items.append(
            {
                "item_id": item_id,
                "category": _text(row["category"]),
                "name": _text(row["name"]) or item_id,
                "prompt_template": prompt,
                "output_filename": output_filename,
                "aspect_ratio": aspect_ratio,
                "image_size": image_size,
                "enabled": _enabled(row["enabled"]),
                "reference_role": _text(row["reference_role"]),
                "notes": _text(row["notes"]),
            }
        )

    if not items:
        raise ValueError("Workbook contains no prompt rows")

    resolved_pack_id = _slug(pack_id or source.parent.name or source.stem)
    destination = Path(output_path).resolve() if output_path else source.with_name("pack.json")
    destination.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema_version": 1,
        "pack_id": resolved_pack_id,
        "name": resolved_pack_id,
        "source_workbook": source.name,
        "compiled_at": datetime.now().isoformat(timespec="seconds"),
        "item_count": len(items),
        "enabled_count": sum(1 for item in items if item["enabled"]),
        "items": items,
    }
    destination.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return destination


def create_placeholder_workbook(output_path: str | Path, *, pack_id: str) -> Path:
    """Create a small editable placeholder pack for later prompt expansion."""
    openpyxl = _require_openpyxl()
    destination = Path(output_path).resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "prompts"
    sheet.append(REQUIRED_COLUMNS)
    rows = [
        (
            "massing-analysis",
            "形体分析",
            "建筑体量生成逻辑",
            "Use all supplied reference renderings as the same project. Create a clean architectural massing-analysis diagram that explains the major volumes, additions, subtractions, and hierarchy. Preserve the building identity from the references. No labels or invented measurements.",
            "analysis_massing.png",
            "16:9",
            "1K",
            True,
            "project-render",
            "Placeholder prompt; replace with the final architecture pack wording.",
        ),
        (
            "circulation-analysis",
            "流线分析",
            "概念流线示意",
            "Use all supplied reference renderings as the same project. Create an architectural circulation-analysis illustration with restrained diagrammatic paths and clear entry, approach, and movement hierarchy inferred from the visible design. Preserve the project form. No text labels or measurements.",
            "analysis_circulation.png",
            "16:9",
            "1K",
            True,
            "project-render",
            "Placeholder prompt; runs even when only one rendering is supplied.",
        ),
        (
            "environment-analysis",
            "环境分析",
            "日照与环境关系示意",
            "Use all supplied reference renderings as the same project. Create a polished architectural environmental-analysis diagram showing conceptual sun, wind, landscape, and context relationships around the visible building. Keep the building recognizable and the diagram visually presentation-ready. No text labels.",
            "analysis_environment.png",
            "16:9",
            "1K",
            True,
            "project-render",
            "Placeholder prompt; replace when the complete Excel library is ready.",
        ),
    ]
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    widths = {"A": 24, "B": 18, "C": 28, "D": 90, "E": 32, "F": 14, "G": 12, "H": 12, "I": 22, "J": 52}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    metadata = workbook.create_sheet("metadata")
    metadata.append(("pack_id", _slug(pack_id)))
    metadata.append(("description", "Placeholder architectural analysis prompt pack"))
    workbook.save(destination)
    return destination


def build_image_manifest(
    pack_path: str | Path,
    project_path: str | Path,
    reference_images: list[str],
    *,
    provider_profile: str = DEFAULT_PROFILE,
    prompt_context: str = "",
) -> Path:
    """Build image_prompts.json for all enabled rows in a compiled pack."""
    pack_file = Path(pack_path).resolve()
    pack = json.loads(pack_file.read_text(encoding="utf-8"))
    project = Path(project_path).resolve()
    images_dir = project / "images"
    images_dir.mkdir(parents=True, exist_ok=True)
    resolved_references: list[str] = []
    for value in reference_images:
        candidate = Path(value)
        if not candidate.is_absolute():
            candidate = images_dir / candidate
        candidate = candidate.resolve()
        if not candidate.is_file():
            raise ValueError(f"Reference image not found: {candidate}")
        try:
            resolved_references.append(candidate.relative_to(images_dir).as_posix())
        except ValueError:
            resolved_references.append(str(candidate))

    items = []
    for item in pack.get("items", []):
        if not item.get("enabled"):
            continue
        prompt = item["prompt_template"]
        if prompt_context.strip():
            prompt = f"{prompt_context.strip()}\n\n{prompt}"
        items.append(
            {
                "filename": item["output_filename"],
                "purpose": item.get("name") or item["item_id"],
                "page_role": "local",
                "text_policy": "none",
                "aspect_ratio": item["aspect_ratio"],
                "image_size": item["image_size"],
                "prompt": prompt,
                "status": "Pending",
                "reference_images": resolved_references,
                "provider_profile": provider_profile,
                "analysis_pack_id": pack["pack_id"],
                "analysis_item_id": item["item_id"],
            }
        )
    if not items:
        raise ValueError(f"No enabled items in pack: {pack_file}")

    manifest = {
        "schema_version": 2,
        "project": project.name,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "provider_profile": provider_profile,
        "analysis_pack_id": pack["pack_id"],
        "items": items,
    }
    destination = images_dir / "image_prompts.json"
    destination.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return destination


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Manage Excel analysis prompt packs.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_parser = subparsers.add_parser("init", help="Create a placeholder Excel pack")
    init_parser.add_argument("output", help="Output .xlsx path")
    init_parser.add_argument("--pack-id", required=True, help="Pack id")

    compile_parser = subparsers.add_parser("compile", help="Validate and compile an Excel pack")
    compile_parser.add_argument("workbook", help="Input .xlsx path")
    compile_parser.add_argument("-o", "--output", help="Output pack.json path")
    compile_parser.add_argument("--pack-id", help="Override pack id")

    manifest_parser = subparsers.add_parser("build-manifest", help="Build image_prompts.json")
    manifest_parser.add_argument("pack", help="Compiled pack.json")
    manifest_parser.add_argument("project", help="PPT Master project path")
    manifest_parser.add_argument("--reference", action="append", default=[], help="Reference image")
    manifest_parser.add_argument("--profile", default=DEFAULT_PROFILE, help="Provider profile")
    manifest_parser.add_argument("--prompt-context", default="", help="Project context prepended to prompts")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        if args.command == "init":
            output = create_placeholder_workbook(args.output, pack_id=args.pack_id)
        elif args.command == "compile":
            output = compile_workbook(args.workbook, args.output, pack_id=args.pack_id)
        else:
            output = build_image_manifest(
                args.pack,
                args.project,
                args.reference,
                provider_profile=args.profile,
                prompt_context=args.prompt_context,
            )
    except (OSError, ValueError, RuntimeError, json.JSONDecodeError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
